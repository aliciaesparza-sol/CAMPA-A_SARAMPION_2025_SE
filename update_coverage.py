import pandas as pd
import os
from openpyxl import load_workbook

csv_path = r"c:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\CAMPAÑA SARAMPIÓN 10 SEMANAS\SRP-SR-2025_22-02-2026 06-41-08.csv"
excel_path = r"C:\Users\aicil\.gemini\antigravity\scratch\coverage_copy.xlsx"

print("Reading CSV...")
df_csv = pd.read_csv(csv_path, encoding='latin1', low_memory=False)

# Robustly find dose columns for 10-49 age groups
all_cols = df_csv.columns.tolist()
target_keywords = ['10 A 12', '13 A 19', '20 A 39', '40 A 49']
dose_cols = [c for c in all_cols if any(k in c for k in target_keywords)]

print(f"Matched columns: {dose_cols}")

# Ensure numeric and fill NaNs
for col in dose_cols:
    df_csv[col] = pd.to_numeric(df_csv[col], errors='coerce').fillna(0)

df_csv['Total_10_49'] = df_csv[dose_cols].sum(axis=1)

agg_doses = df_csv.groupby('MUNICIPIO')['Total_10_49'].sum().reset_index()
agg_doses['MUNICIPIO'] = agg_doses['MUNICIPIO'].str.upper().str.strip()

print("--- Aggregated Doses (Sample) ---")
print(agg_doses.head(10))

print("Updating Excel...")
wb = load_workbook(excel_path)
ws = wb.active # Assuming the first sheet

# In the Excel file seen earlier:
# Column 2 (B) seems to be Municipality Name (Unnamed: 1)
# Column 4 (D) seems to be Doses (Unnamed: 3)
muni_col = 2
dose_col_idx = 4

updated_count = 0
for row in range(1, ws.max_row + 1):
    cell_val = ws.cell(row=row, column=muni_col).value
    if cell_val and isinstance(cell_val, str):
        muni_name = cell_val.upper().strip()
        match = agg_doses[agg_doses['MUNICIPIO'] == muni_name]
        if not match.empty:
            total_val = match['Total_10_49'].values[0]
            ws.cell(row=row, column=dose_col_idx).value = total_val
            print(f"Updated {muni_name} at row {row} with {total_val}")
            updated_count += 1

final_path = r"C:\Users\aicil\.gemini\antigravity\scratch\COBERTURAS_UPDATED_2025.xlsx"
wb.save(final_path)
print(f"\nSuccessfully updated {updated_count} municipalities.")
print(f"File saved to: {final_path}")

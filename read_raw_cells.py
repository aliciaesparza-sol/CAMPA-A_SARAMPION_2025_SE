import openpyxl
import sys

def read_raw():
    wb = openpyxl.load_workbook('data.xlsx', read_only=True, data_only=True)
    
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"\n{'='*60}")
        print(f"SHEET: {sn}")
        print(f"{'='*60}")
        
        rows = list(ws.iter_rows(min_row=1, max_row=8, values_only=True))
        for r_idx, row in enumerate(rows):
            non_empty = [(c_idx, v) for c_idx, v in enumerate(row) if v is not None]
            if non_empty:
                print(f"  Row {r_idx+1}: {non_empty}")

    wb.close()

if __name__ == "__main__":
    read_raw()

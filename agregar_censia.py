
# -*- coding: utf-8 -*-
"""
Agrega al Excel de cobertura:
  H: Dosis CENSIA (del Anexo 1)
  I: Cobertura CENSIA (%)
  J: Avance CENSIA (barra DataBar)
Con semafarizacion y fuente al pie.
"""

import openpyxl, glob
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule

# ─── RUTAS ────────────────────────────────────────────────────────────────────
files = glob.glob(r'c:\Users\aicil\OneDrive\ANTIGRAVITY\Anexo*.xlsx')
CENSIA_PATH = files[0] if files else None
DEST_IN  = r'c:\Users\aicil\OneDrive\ANTIGRAVITY\Cobertura_tmp_src.xlsx'
DEST     = r'c:\Users\aicil\OneDrive\ANTIGRAVITY\Cobertura_SRP-SR_SSA_Durango_2026_CENSIA.xlsx'
URL_FUENTE = 'https://siscensia.salud.gob.mx/sarampion_2025/ssa/reporte.php'

if not CENSIA_PATH:
    raise FileNotFoundError("No se encontro el archivo Anexo 1 de CENSIA")
print("Leyendo CENSIA:", CENSIA_PATH)

# ─── EXTRAER DATOS DURANGO DE CENSIA ─────────────────────────────────────────
wb_c = openpyxl.load_workbook(CENSIA_PATH, data_only=True)

# --- Poblacion META Durango ---
ws_meta = wb_c['Pob META']
meta_censia = {}
for row in ws_meta.iter_rows(min_row=3, max_row=35, values_only=True):
    if row[0] and 'DURANGO' in str(row[0]).upper():
        # cols: Grupo | 6-11m | 1anio | 18m | Rez2-12 | 13-19 | 20-39 | 40-49 | TOTAL
        meta_censia['6 a 11 meses']        = row[1] or 0
        meta_censia['1 anio']              = row[2] or 0
        meta_censia['18 meses']            = row[3] or 0
        meta_censia['Rezagados 2-12']      = row[4] or 0
        meta_censia['13 a 19']             = row[5] or 0
        meta_censia['20 a 39']             = row[6] or 0
        meta_censia['40 a 49']             = row[7] or 0
        break

# --- Dosis aplicadas Durango ---
ws_dos = wb_c['Dosis aplicadas']
dosis_censia = {}
for row in ws_dos.iter_rows(min_row=3, max_row=35, values_only=True):
    if row[0] and 'DURANGO' in str(row[0]).upper():
        dosis_censia['6 a 11 meses']   = int(row[1] or 0)
        dosis_censia['1 anio']         = int(row[2] or 0)
        dosis_censia['18 meses']       = int(row[3] or 0)
        dosis_censia['Rezagados 2-12'] = int(row[4] or 0)
        dosis_censia['13 a 19']        = int(row[5] or 0)
        dosis_censia['20 a 39']        = int(row[6] or 0)
        dosis_censia['40 a 49']        = int(row[7] or 0)
        break

print("Poblacion meta CENSIA Durango:", meta_censia)
print("Dosis CENSIA Durango:", dosis_censia)

# ─── ESTILOS ──────────────────────────────────────────────────────────────────
thin       = Side(style='thin', color='BFBFBF')
border     = Border(top=thin, left=thin, right=thin, bottom=thin)
HEADER_FILL = PatternFill(fgColor='8B1A4A', fill_type='solid')   # vino oscuro
HEADER_FONT = Font(color='FFFFFF', bold=True, size=10)
RED_FILL    = PatternFill(fgColor='FFB3B3', fill_type='solid')
ORANGE_FILL = PatternFill(fgColor='FFD9A0', fill_type='solid')
GREEN_FILL  = PatternFill(fgColor='B7E4C7', fill_type='solid')
TOTAL_FILL  = PatternFill(fgColor='D6C5E0', fill_type='solid')  # morado claro

# ─── ABRIR EXCEL DE COBERTURA SSA ─────────────────────────────────────────────
wb = openpyxl.load_workbook(DEST_IN)
ws = wb.active

# Orden: A=Grupo B=Pob2026 C=Meta D=Dosis E=Susceptible F=Cobertura% G=Avance
# Nuevo: H=DosasCENSIA I=CoberturaCENSIA% J=AvanceCENSIA

# Widths
ws.column_dimensions['H'].width = 18
ws.column_dimensions['I'].width = 16
ws.column_dimensions['J'].width = 18

# ─── ENCABEZADOS FILA 3 ───────────────────────────────────────────────────────
headers_censia = {
    8: 'Dosis CENSIA',
    9: 'Cobertura CENSIA (%)',
    10: 'Avance CENSIA',
}
for col, title in headers_censia.items():
    c = ws.cell(row=3, column=col)
    c.value      = title
    c.font       = HEADER_FONT
    c.fill       = HEADER_FILL
    c.alignment  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border     = border

# ─── MAPEO: fila -> clave CENSIA ──────────────────────────────────────────────
fila_clave = {
    4:  '6 a 11 meses',
    5:  '1 anio',
    6:  '18 meses',
    7:  'Rezagados 2-12',
    8:  '13 a 19',
    9:  '20 a 39',
    10: '40 a 49',
}

total_dosis_c  = 0
total_meta_c   = 0

for fila, clave in fila_clave.items():
    dosis = dosis_censia.get(clave, 0)
    meta  = meta_censia.get(clave, 0)
    total_dosis_c += dosis
    total_meta_c  += meta

    cob = round((dosis / meta) * 100, 2) if meta else 0.0
    avance = min(cob / 100.0, 1.0)

    # H: Dosis CENSIA
    ch = ws.cell(row=fila, column=8)
    ch.value = dosis
    ch.number_format = '#,##0'
    ch.alignment = Alignment(horizontal='center', vertical='center')
    ch.border = border

    # I: Cobertura %
    ci = ws.cell(row=fila, column=9)
    ci.value = cob
    ci.number_format = '#,##0.00"%"'
    ci.alignment = Alignment(horizontal='center', vertical='center')
    ci.border = border

    # J: Avance (decimal para DataBar)
    cj = ws.cell(row=fila, column=10)
    cj.value = avance
    cj.number_format = '0.0%'
    cj.alignment = Alignment(horizontal='center', vertical='center')
    cj.border = border

    # Colorear I y J segun semaforo
    if cob < 50:
        fill = RED_FILL
    elif cob < 80:
        fill = ORANGE_FILL
    else:
        fill = GREEN_FILL
    ci.fill = fill
    cj.fill = fill

# ─── FILA TOTAL (11) ──────────────────────────────────────────────────────────
total_cob_c = round((total_dosis_c / total_meta_c) * 100, 2) if total_meta_c else 0.0
total_avance_c = min(total_cob_c / 100.0, 1.0)

for col, val, fmt in [
    (8,  total_dosis_c,   '#,##0'),
    (9,  total_cob_c,     '#,##0.00"%"'),
    (10, total_avance_c,  '0.0%'),
]:
    c = ws.cell(row=11, column=col)
    c.value = val
    c.number_format = fmt
    c.fill = TOTAL_FILL
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border

# ─── DATA BAR en columna J ────────────────────────────────────────────────────
db_rule = DataBarRule(
    start_type='num', start_value=0,
    end_type='num',   end_value=1,
    color='8B1A4A'   # vino (para diferenciar del azul de SSA)
)
ws.conditional_formatting.add('J4:J11', db_rule)

# ─── FUENTE AL PIE ────────────────────────────────────────────────────────────
# Buscar ultima fila con contenido
ultima = 11
for r in range(12, 25):
    if ws.cell(row=r, column=1).value is not None:
        ultima = r

pie_row = ultima + 1
# Buscar donde empieza la leyenda (puede haber contenido antes)
for r in range(pie_row, pie_row + 10):
    if ws.cell(row=r, column=1).value is not None:
        pie_row = r + 1

# Fuente SSA
f1 = ws.cell(row=pie_row, column=1)
f1.value = 'Fuente SSA: SRP-SR-2025_22-02-2026 06-41-08.csv'
f1.font  = Font(italic=True, size=8, color='444444')

# Fuente CENSIA
f2 = ws.cell(row=pie_row + 1, column=1)
f2.value = f'Fuente CENSIA: {URL_FUENTE}'
f2.font  = Font(italic=True, size=8, color='444444', underline='single')
ws.merge_cells(start_row=pie_row+1, start_column=1, end_row=pie_row+1, end_column=10)

# Fecha del archivo CENSIA
ws.cell(row=pie_row + 2, column=1).value = 'Datos CENSIA: RDA 3er Trimestre 2025 + SIS CENSIA 1 Oct 2025 al 6 Feb 2026'
ws.cell(row=pie_row + 2, column=1).font  = Font(italic=True, size=8, color='444444')

wb.save(DEST)
print("\nExcel actualizado con columnas CENSIA. Guardado en:")
print(DEST)

print("\n=== CENSIA DURANGO ===")
keys = ['6 a 11 meses','1 anio','18 meses','Rezagados 2-12','13 a 19','20 a 39','40 a 49']
labels = ['6-11 meses','1 anio','18 meses','Rezag.2-12','13-19','20-39','40-49']
for k, lbl in zip(keys, labels):
    d = dosis_censia.get(k,0)
    m = meta_censia.get(k,0)
    c = round(d/m*100,2) if m else 0
    print(f"  {lbl:<15} Dosis: {d:>7,}  Meta: {m:>7,}  Cob: {c:>6.2f}%")
print(f"  {'TOTAL':<15} Dosis: {total_dosis_c:>7,}  Meta: {total_meta_c:>7,}  Cob: {total_cob_c:>6.2f}%")

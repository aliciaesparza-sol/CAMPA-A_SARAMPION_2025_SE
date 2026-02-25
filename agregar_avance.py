
# -*- coding: utf-8 -*-
"""
Agrega columna de % de Avance de Cobertura al Excel,
con DataBar (barra de progreso nativa de Excel) y formato condicional
semaforo: rojo < 50%, naranja 50-80%, verde >= 80%.
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule, Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter

SRC  = r'c:\Users\aicil\OneDrive\ANTIGRAVITY\Cobertura_SRP-SR_SSA_Durango_2026_actualizado.xlsx'
DEST = r'c:\Users\aicil\OneDrive\ANTIGRAVITY\Cobertura_SRP-SR_SSA_Durango_2026_actualizado.xlsx'

wb = openpyxl.load_workbook(SRC)
ws = wb.active

# Estilos de borde
thin = Side(style='thin', color='BFBFBF')
border = Border(top=thin, left=thin, right=thin, bottom=thin)

# Colores de semaforo
RED_FILL    = PatternFill(fgColor='FFB3B3', fill_type='solid')   # rojo claro < 50%
ORANGE_FILL = PatternFill(fgColor='FFD9A0', fill_type='solid')   # naranja 50-79%
GREEN_FILL  = PatternFill(fgColor='B7E4C7', fill_type='solid')   # verde >= 80%
TOTAL_FILL  = PatternFill(fgColor='D6E4F7', fill_type='solid')   # azul para total

HEADER_FILL = PatternFill(fgColor='1F3864', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=10)

# Filas de datos (4 a 10) y total (11)
DATA_ROWS  = range(4, 11)
TOTAL_ROW  = 11

# ─── AGREGAR ENCABEZADO EN COLUMNA G ────────────────────────────────────────
ws.column_dimensions['G'].width = 22

hdr = ws.cell(row=3, column=7)
hdr.value        = 'Avance Cobertura'
hdr.font         = HEADER_FONT
hdr.fill         = HEADER_FILL
hdr.alignment    = Alignment(horizontal='center', vertical='center', wrap_text=True)
hdr.border       = border

# ─── LLENAR VALORES EN G (igual a cobertura en F / 100) ─────────────────────
for row_num in list(DATA_ROWS) + [TOTAL_ROW]:
    cob_val = ws.cell(row=row_num, column=6).value or 0.0
    # Limitar al 100% para barra de progreso
    avance  = min(float(cob_val) / 100.0, 1.0)

    cell = ws.cell(row=row_num, column=7)
    cell.value     = avance
    cell.number_format = '0.0%'
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border    = border

    # colorear la celda de cobertura (F) segun semaforo
    cob_cell = ws.cell(row=row_num, column=6)
    cob_cell.number_format = '#,##0.00"%"'
    cob_cell.alignment = Alignment(horizontal='center', vertical='center')

    if row_num == TOTAL_ROW:
        cell.fill     = TOTAL_FILL
        cell.font     = Font(bold=True)
        cob_cell.fill = TOTAL_FILL
        cob_cell.font = Font(bold=True)
    elif cob_val < 50:
        cell.fill     = RED_FILL
        cob_cell.fill = RED_FILL
    elif cob_val < 80:
        cell.fill     = ORANGE_FILL
        cob_cell.fill = ORANGE_FILL
    else:
        cell.fill     = GREEN_FILL
        cob_cell.fill = GREEN_FILL

# ─── AGREGAR DATA BAR (barra de progreso nativa de Excel) ────────────────────
# Aplica sobre G4:G10
data_range = f'G4:G{TOTAL_ROW}'

db_rule = DataBarRule(
    start_type='num', start_value=0,
    end_type='num',   end_value=1,
    color='2196F3'   # azul material design
)
ws.conditional_formatting.add(data_range, db_rule)

# ─── COLUMNA F: tambien DataBar semaforo de color ────────────────────────────
# No se puede mezclar DataBar con fill de celda en todas las versiones,
# pero agregamos una escala de color para que se vea el gradiente
cob_range = f'F4:F{TOTAL_ROW - 1}'   # excluir total del formato

# Escala de color: rojo (0) -> amarillo (50) -> verde (100)
cs_rule = ColorScaleRule(
    start_type='num',  start_value=0,   start_color='FF4444',
    mid_type='num',    mid_value=50,    mid_color='FFA500',
    end_type='num',    end_value=100,   end_color='00AA00'
)
ws.conditional_formatting.add(cob_range, cs_rule)

# ─── AJUSTAR anchos de columnas clave ────────────────────────────────────────
ws.column_dimensions['A'].width = 26
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 13
ws.column_dimensions['G'].width = 18

# Forzar altura de fila del header
ws.row_dimensions[3].height = 30

# ─── LEYENDA al pie ──────────────────────────────────────────────────────────
leyenda_row = TOTAL_ROW + 2
ws.cell(row=leyenda_row, column=1).value = 'Semaforización:'
ws.cell(row=leyenda_row, column=1).font  = Font(bold=True, size=9)

colores = [
    (RED_FILL,    'Rojo: < 50% cobertura'),
    (ORANGE_FILL, 'Naranja: 50% - 79%'),
    (GREEN_FILL,  'Verde: >= 80%'),
]
for idx, (fill, txt) in enumerate(colores):
    col_lbl  = leyenda_row + idx + 1
    ic = ws.cell(row=col_lbl, column=1)
    ic.value = txt
    ic.fill  = fill
    ic.font  = Font(size=9)
    ic.alignment = Alignment(indent=1)

wb.save(DEST)
print('Excel actualizado con barra de avance guardado en:')
print(DEST)

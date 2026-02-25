"""
Final script: Extract SRP-SR SSA doses for Durango from CSV, 
combine with Population 2026 and Meta from the original Excel file,
and calculate vaccination coverage by age group.
"""
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ─── 1. Read CSV (already filtered for Durango SSA) ─────────────────────────
df_csv = pd.read_csv('doses.csv', encoding='latin-1', sep=',', engine='python')
df_csv.columns = [str(c).strip() for c in df_csv.columns]

print("Total CSV rows:", len(df_csv))
print("CSV columns:")
for i, c in enumerate(df_csv.columns):
    print(f"  [{i}] {c}")

# ─── 2. Identify SRP-SR columns ──────────────────────────────────────────────
# The CSV uses SRP for vaccine type, with age group in the column name.
# We need to sum all SRP* + SR* columns that correspond to the requested age groups.

# Age group mapping: user label -> keywords to find in CSV column names
# Looking at: 'SRP 6 A 11 MESES', 'SRP 1 A?O', 'SRP 18 MESES', 
#             'SRP 2 A 12 A?OS' / 'SRP REZAGADOS', 'SRP 13 A 19', 
#             'SRP 20 A 39', 'SRP 40 A 49'
AGE_GROUPS = {
    '6 a 11 meses':          ['6 A 11', '6A11', '6-11'],
    '1 año':                 ['1 A', '1A ', '12 MES', '1 AN', 'MENORES'],
    '18 meses':              ['18 MES', '18M'],
    'Rezagados 2 a 12 años': ['2 A 12', '2A12', 'REZAG', '2-12'],
    '13 a 19 años':          ['13 A 19', '13A19', '13-19'],
    '20 a 39 años':          ['20 A 39', '20A39', '20-39'],
    '40 a 49 años':          ['40 A 49', '40A49', '40-49'],
}

def find_matching_cols(all_cols, keywords):
    """Find columns that match any keyword (case-insensitive)."""
    matched = []
    for col in all_cols:
        col_up = col.upper()
        if any(kw.upper() in col_up for kw in keywords):
            # Only include SRP or SR columns (vaccine specific)
            if any(x in col_up for x in ['SRP', ' SR ']):
                matched.append(col)
    return matched

col_map = {}
for group, kws in AGE_GROUPS.items():
    matched = find_matching_cols(df_csv.columns, kws)
    col_map[group] = matched
    print(f"\n[{group}] matched cols: {matched}")

# ─── 3. Sum doses per age group ──────────────────────────────────────────────
group_doses = {}
for group, cols in col_map.items():
    if cols:
        total = df_csv[cols].apply(pd.to_numeric, errors='coerce').fillna(0).sum().sum()
        group_doses[group] = int(total)
    else:
        group_doses[group] = 0

print("\nDoses per age group:")
for g, d in group_doses.items():
    print(f"  {g}: {d}")

# ─── 4. Get Population and Meta from original data.xlsx ─────────────────────
src = 'data.xlsx'
xl = pd.ExcelFile(src)

def load_sheet(sn):
    df_raw = pd.read_excel(src, sheet_name=sn, header=None, nrows=10)
    best_row = max(range(min(8, len(df_raw))),
                   key=lambda r: df_raw.iloc[r].apply(
                       lambda x: isinstance(x, str) and len(str(x).strip()) > 1).sum())
    df = pd.read_excel(src, sheet_name=sn, header=best_row)
    df.columns = [str(c).strip() for c in df.columns]
    return df

def get_durango(df):
    mask = df.apply(lambda c: c.astype(str).str.contains('DURANGO', case=False, na=False)).any(axis=1)
    rows = df[mask]
    return rows.iloc[0] if not rows.empty else None

# Age group column keywords for population/meta sheets
POP_COL_KW = {
    '6 a 11 meses':          ['6 a 11', '6-11', '6 mes'],
    '1 año':                 ['1 año', '1año', '1 a\u00f1', '< 1', 'menor 1'],
    '18 meses':              ['18 mes', '18m'],
    'Rezagados 2 a 12 años': ['rezag', '2 a 12', '2-12'],
    '13 a 19 años':          ['13 a 19', '13-19'],
    '20 a 39 años':          ['20 a 39', '20-39'],
    '40 a 49 años':          ['40 a 49', '40-49'],
}

def find_col(columns, keywords):
    for col in columns:
        if any(kw.lower() in col.lower() for kw in keywords):
            return col
    return None

def safe_num(val):
    try:
        v = float(val)
        return 0 if pd.isna(v) else v
    except:
        return 0

# Find sheet names
pob_sn  = next((s for s in xl.sheet_names if 'Total' in s or 'total' in s.lower() or 'Poblaci' in s), xl.sheet_names[0])
meta_sn = next((s for s in xl.sheet_names if 'META' in s.upper() or 'meta' in s.lower()), None)

df_pob  = load_sheet(pob_sn)
df_meta = load_sheet(meta_sn) if meta_sn else None

dur_pob  = get_durango(df_pob)
dur_meta = get_durango(df_meta) if df_meta is not None else None

rows = []
for group in AGE_GROUPS.keys():
    kws = POP_COL_KW[group]
    
    pob_col  = find_col(df_pob.columns,  kws) if dur_pob  is not None else None
    meta_col = find_col(df_meta.columns, kws) if dur_meta is not None else None
    
    pob  = safe_num(dur_pob[pob_col])   if pob_col  and pob_col  in dur_pob  else 0
    meta = safe_num(dur_meta[meta_col]) if meta_col and meta_col in dur_meta else 0
    apl  = group_doses.get(group, 0)
    sus  = max(0, meta - apl)
    cob  = round(apl / meta * 100, 2) if meta > 0 else 0
    
    rows.append({
        'Grupo de Edad':         group,
        'Población 2026':        int(pob),
        'Población Meta':        int(meta),
        'Dosis Aplicadas (CSV)': int(apl),
        'Población Susceptible': int(sus),
        'Cobertura (%)':         cob,
    })

print("\nFINAL TABLE:")
df_out = pd.DataFrame(rows)
print(df_out.to_string(index=False))

# ─── 5. Write to Excel ───────────────────────────────────────────────────────
OUT = r'C:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\CAMPAÑA SARAMPIÓN 10 SEMANAS\Cobertura_SRP-SR_SSA_Durango_2026.xlsx'
os.makedirs(os.path.dirname(OUT), exist_ok=True)

wb = Workbook()
ws = wb.active
ws.title = "Cobertura SRP-SR SSA Durango"

# Styles
hdr_fill  = PatternFill("solid", fgColor="1F4E79")
hdr_font  = Font(bold=True, color="FFFFFF", size=11)
sub_fill  = PatternFill("solid", fgColor="2E75B6")
sub_font  = Font(bold=True, color="FFFFFF", size=10)
alt_fill  = PatternFill("solid", fgColor="DEEAF1")
wht_fill  = PatternFill("solid", fgColor="FFFFFF")
grn_fill  = PatternFill("solid", fgColor="C6EFCE")
yel_fill  = PatternFill("solid", fgColor="FFEB9C")
red_fill  = PatternFill("solid", fgColor="FFC7CE")
center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_al   = Alignment(horizontal="left",   vertical="center")
thin      = Side(style="thin", color="9DC3E6")
bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)

# Title
ws.merge_cells("A1:F1")
ws["A1"] = "COBERTURA DE VACUNACIÓN SRP-SR SSA — DURANGO 2026"
ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
ws["A1"].fill = hdr_fill
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36

# Source note
ws.merge_cells("A2:F2")
ws["A2"] = f"Fuente dosis aplicadas: SRP-SR-2025_22-02-2026.csv (SRP-SR SSA)  |  Metodología: Cobertura (%) = Dosis Aplicadas / Población Meta × 100"
ws["A2"].font = Font(italic=True, size=9, color="1F4E79")
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 20

# Headers
headers = ['Grupo de Edad', 'Población 2026', 'Población Meta', 'Dosis Aplicadas (CSV)', 'Población Susceptible', 'Cobertura (%)']
for c_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=3, column=c_idx, value=h)
    cell.font = sub_font
    cell.fill = sub_fill
    cell.alignment = center
    cell.border = bdr
ws.row_dimensions[3].height = 30

# Data rows
for r_idx, row in enumerate(rows, 4):
    fill = alt_fill if r_idx % 2 == 0 else wht_fill
    for c_idx, col_name in enumerate(headers, 1):
        val = row[col_name]
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.border = bdr
        if col_name == 'Grupo de Edad':
            cell.alignment = left_al
            cell.font = Font(bold=True, size=10)
            cell.fill = fill
        elif col_name == 'Cobertura (%)':
            cell.alignment = center
            cell.font = Font(bold=True, size=10)
            cell.number_format = '0.00"%"'
            if val >= 95:
                cell.fill = grn_fill
            elif val >= 80:
                cell.fill = yel_fill
            else:
                cell.fill = red_fill
        else:
            cell.alignment = center
            cell.number_format = '#,##0'
            cell.fill = fill
    ws.row_dimensions[r_idx].height = 22

# Totals
tr = len(rows) + 4
ws.cell(row=tr, column=1, value="TOTAL DURANGO").font = Font(bold=True, color="FFFFFF")
ws.cell(row=tr, column=1).fill = hdr_fill
ws.cell(row=tr, column=1).alignment = left_al
ws.cell(row=tr, column=1).border = bdr

for c_idx, col_name in enumerate(headers[1:], 2):
    if col_name == 'Cobertura (%)':
        t_meta = sum(r['Población Meta'] for r in rows)
        t_apl  = sum(r['Dosis Aplicadas (CSV)'] for r in rows)
        val = round(t_apl / t_meta * 100, 2) if t_meta > 0 else 0
        cell = ws.cell(row=tr, column=c_idx, value=val)
        cell.number_format = '0.00"%"'
        if val >= 95: cell.fill = grn_fill
        elif val >= 80: cell.fill = yel_fill
        else: cell.fill = red_fill
        cell.font = Font(bold=True)
    else:
        val = sum(r[col_name] for r in rows)
        cell = ws.cell(row=tr, column=c_idx, value=val)
        cell.number_format = '#,##0'
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hdr_fill
    cell.alignment = center
    cell.border = bdr
ws.row_dimensions[tr].height = 25

# Legend
lr = tr + 2
ws.merge_cells(f"A{lr}:F{lr}")
ws.cell(row=lr, column=1).value = "🟢 ≥ 95% Buena  |  🟡 80-94% Regular  |  🔴 < 80% Insuficiente"
ws.cell(row=lr, column=1).font = Font(italic=True, size=9)
ws.cell(row=lr, column=1).alignment = Alignment(horizontal="left")

# Column widths
ws.column_dimensions['A'].width = 28
for col in ['B','C','D','E','F']:
    ws.column_dimensions[col].width = 22

wb.save(OUT)
print(f"\nArchivo guardado en: {OUT}")

"""
Builds the complete age-group vaccination coverage report for DURANGO.
Reads the four sheets, finds the best matching columns for each age group,
then creates a nicely formatted Excel file.
"""
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json

SRC = 'data.xlsx'
OUT = r'C:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\CAMPAÑA SARAMPIÓN 10 SEMANAS\Cobertura_Vacunacion_Durango_2026.xlsx'

SHEETS = {
    'pob':  'PoblaciónTotal 2026',
    'meta': 'Pob META',
    'apl':  'Dosis aplicadas',
    'sus':  'Susceptibles',
}

# Age groups that the user wants
AGE_GROUPS = [
    '6 a 11 meses',
    '1 año',
    '18 meses',
    'Rezagados 2 a 12 años',
    '13 a 19 años',
    '20 a 39 años',
    '40 a 49 años',
]

def load_sheet(sn):
    """Load a sheet, auto-detect header row, return cleaned DataFrame."""
    df_raw = pd.read_excel(SRC, sheet_name=sn, header=None, nrows=10)
    best_row = max(range(min(8, len(df_raw))),
                   key=lambda r: df_raw.iloc[r].apply(
                       lambda x: isinstance(x, str) and len(str(x).strip()) > 1).sum())
    df = pd.read_excel(SRC, sheet_name=sn, header=best_row)
    df.columns = [str(c).strip() for c in df.columns]
    return df

def get_durango(df):
    """Return the Durango row as a Series."""
    mask = df.apply(lambda c: c.astype(str).str.contains('DURANGO', case=False, na=False)).any(axis=1)
    rows = df[mask]
    return rows.iloc[0] if not rows.empty else None

def find_col(columns, keywords):
    """Find the best column name that matches any keyword (case-insensitive), return it or None."""
    for col in columns:
        col_lower = col.lower()
        if any(kw.lower() in col_lower for kw in keywords):
            return col
    return None

def safe_num(val):
    """Convert to float, return 0 if non-numeric."""
    try:
        v = float(val)
        return 0 if pd.isna(v) else v
    except:
        return 0

# Load all sheets
dfs = {key: load_sheet(sn) for key, sn in SHEETS.items()}

# Print all columns for debugging — saved to JSON
col_info = {key: list(df.columns) for key, df in dfs.items()}
with open('col_info_final.json', 'w', encoding='utf-8') as f:
    json.dump(col_info, f, ensure_ascii=False, indent=2)

# Keywords for each age group column (order matters — first match wins)
AGE_COL_KEYWORDS = {
    '6 a 11 meses':         ['6 a 11', '6-11', '6 meses', '6a11'],
    '1 año':                ['1 año', '1 a\u00f1o', '1año', '12 mes', '1 a?o', '1a\u00f1', '< 1', 'menor 1', 'menores'],
    '18 meses':             ['18 mes', '18m', '18\u00a0mes'],
    'Rezagados 2 a 12 años':['rezag', '2 a 12', '2-12'],
    '13 a 19 años':         ['13 a 19', '13-19', '13 a\u00f1'],
    '20 a 39 años':         ['20 a 39', '20-39', '20 a\u00f1'],
    '40 a 49 años':         ['40 a 49', '40-49', '40 a\u00f1'],
}

# Get Durango rows
dur = {key: get_durango(df) for key, df in dfs.items()}

# Build mapping: for each sheet and each age group, what column is it?
col_map = {}
for key, df in dfs.items():
    col_map[key] = {}
    for group, kws in AGE_COL_KEYWORDS.items():
        col_name = find_col(df.columns, kws)
        col_map[key][group] = col_name

print("\nColumn mapping:")
for key, mapping in col_map.items():
    print(f"\n  Sheet [{key}]:")
    for grp, col in mapping.items():
        val = safe_num(dur[key][col]) if dur[key] is not None and col and col in dur[key] else 'NOT FOUND'
        print(f"    {grp} => col='{col}' val={val}")

# ─── Build the result table ───────────────────────────────────────────────────
cols_in_rows = []
for group in AGE_GROUPS:
    pob_col  = col_map['pob'].get(group)
    meta_col = col_map['meta'].get(group)
    apl_col  = col_map['apl'].get(group)
    sus_col  = col_map['sus'].get(group)
    
    pob  = safe_num(dur['pob'][pob_col])   if dur['pob']  is not None and pob_col  and pob_col  in dur['pob']  else 0
    meta = safe_num(dur['meta'][meta_col]) if dur['meta'] is not None and meta_col and meta_col in dur['meta'] else 0
    apl  = safe_num(dur['apl'][apl_col])   if dur['apl']  is not None and apl_col  and apl_col  in dur['apl']  else 0
    sus  = safe_num(dur['sus'][sus_col])   if dur['sus']  is not None and sus_col  and sus_col  in dur['sus']  else 0

    # Coverage = Dosis Aplicadas / Población Meta * 100
    cobertura = round((apl / meta * 100), 2) if meta > 0 else 0

    cols_in_rows.append({
        'Grupo de Edad':         group,
        'Población 2026':        int(pob)  if pob  else 0,
        'Población Meta':        int(meta) if meta else 0,
        'Dosis Aplicadas':       int(apl)  if apl  else 0,
        'Población Susceptible': int(sus)  if sus  else 0,
        'Cobertura (%)':         cobertura,
    })

df_out = pd.DataFrame(cols_in_rows)
print("\n\nFINAL TABLE:")
print(df_out.to_string(index=False))

# ─── Write to Excel with formatting ──────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Cobertura Durango 2026"

# Styles
hdr_fill   = PatternFill("solid", fgColor="1F4E79")
hdr_font   = Font(bold=True, color="FFFFFF", size=11)
subh_fill  = PatternFill("solid", fgColor="2E75B6")
subh_font  = Font(bold=True, color="FFFFFF", size=10)
alt_fill   = PatternFill("solid", fgColor="DEEAF1")
pct_good   = PatternFill("solid", fgColor="C6EFCE")  # green for ≥95%
pct_mid    = PatternFill("solid", fgColor="FFEB9C")  # yellow for 80-94%
pct_low    = PatternFill("solid", fgColor="FFC7CE")  # red for <80%
center     = Alignment(horizontal="center", vertical="center", wrap_text=True)
left       = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin       = Side(style="thin", color="9DC3E6")
border     = Border(left=thin, right=thin, top=thin, bottom=thin)

# Title row
ws.merge_cells("A1:F1")
ws["A1"] = "COBERTURA DE VACUNACIÓN CONTRA SARAMPIÓN - DURANGO 2026"
ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 35

# Methodology note
ws.merge_cells("A2:F2")
ws["A2"] = "Metodología: Cobertura (%) = Dosis Aplicadas / Población Meta × 100  |  Susceptibles = Población Meta − Dosis Aplicadas"
ws["A2"].font = Font(italic=True, size=9, color="1F4E79")
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 20

# Headers
headers = ['Grupo de Edad', 'Población 2026', 'Población Meta', 'Dosis Aplicadas', 'Población Susceptible', 'Cobertura (%)']
for c_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=3, column=c_idx, value=h)
    cell.font = subh_font
    cell.fill = subh_fill
    cell.alignment = center
    cell.border = border
ws.row_dimensions[3].height = 30

# Data rows
for r_idx, row in enumerate(cols_in_rows, 4):
    fill = alt_fill if r_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
    for c_idx, col_name in enumerate(headers, 1):
        val = row[col_name]
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.border = border
        if col_name == 'Grupo de Edad':
            cell.alignment = left
            cell.font = Font(bold=True, size=10)
            cell.fill = fill
        elif col_name == 'Cobertura (%)':
            cell.alignment = center
            cell.font = Font(bold=True, size=10)
            cell.number_format = '0.00"%"'
            if val >= 95:
                cell.fill = pct_good
            elif val >= 80:
                cell.fill = pct_mid
            else:
                cell.fill = pct_low
        else:
            cell.alignment = center
            cell.number_format = '#,##0'
            cell.fill = fill
    ws.row_dimensions[r_idx].height = 22

# Total row
total_row = len(cols_in_rows) + 4
ws.cell(row=total_row, column=1, value="TOTAL DURANGO").font = Font(bold=True, size=10, color="FFFFFF")
ws.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor="1F4E79")
ws.cell(row=total_row, column=1).alignment = left
ws.cell(row=total_row, column=1).border = border

for c_idx, col_name in enumerate(headers[1:], 2):
    if col_name == 'Cobertura (%)':
        total_meta = sum(r['Población Meta'] for r in cols_in_rows)
        total_apl  = sum(r['Dosis Aplicadas'] for r in cols_in_rows)
        total_cob  = round(total_apl / total_meta * 100, 2) if total_meta > 0 else 0
        cell = ws.cell(row=total_row, column=c_idx, value=total_cob)
        cell.number_format = '0.00"%"'
        cell.font = Font(bold=True, color="FFFFFF")
        if total_cob >= 95:
            cell.fill = pct_good
        elif total_cob >= 80:
            cell.fill = pct_mid
        else:
            cell.fill = pct_low
    else:
        total_val = sum(r[col_name] for r in cols_in_rows)
        cell = ws.cell(row=total_row, column=c_idx, value=total_val)
        cell.number_format = '#,##0'
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = center
    cell.border = border

ws.row_dimensions[total_row].height = 25

# Column widths
ws.column_dimensions['A'].width = 28
for col in ['B','C','D','E','F']:
    ws.column_dimensions[col].width = 20

# Legend row
legend_row = total_row + 2
ws.merge_cells(f"A{legend_row}:F{legend_row}")
ws.cell(row=legend_row, column=1).value = "Leyenda de cobertura:  🟢 ≥ 95% (Buena)   🟡 80-94% (Regular)   🔴 < 80% (Insuficiente)"
ws.cell(row=legend_row, column=1).font = Font(size=9, italic=True)
ws.cell(row=legend_row, column=1).alignment = Alignment(horizontal="left")

os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print(f"\nArchivo guardado en: {OUT}")

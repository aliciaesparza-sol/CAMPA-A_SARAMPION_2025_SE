
# -*- coding: utf-8 -*-
"""
Script: calcular_cobertura_srp.py
Extrae dosis aplicadas del CSV SRP-SR por grupo de edad
y actualiza el archivo de cobertura Excel.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy
import os
from datetime import datetime

# ─── RUTAS ────────────────────────────────────────────────────────────────────
CSV_PATH  = r'c:\Users\aicil\OneDrive\ANTIGRAVITY\SRP-SR-2025_22-02-2026 06-41-08.csv'
XLSX_SRC  = r'c:\Users\aicil\OneDrive\ANTIGRAVITY\Cobertura_SRP-SR_SSA_Durango_2026.xlsx'
XLSX_OUT  = r'c:\Users\aicil\OneDrive\ANTIGRAVITY\Cobertura_SRP-SR_SSA_Durango_2026_actualizado.xlsx'

# ─── LEER CSV ─────────────────────────────────────────────────────────────────
print("Leyendo CSV...")
df = pd.read_csv(CSV_PATH, low_memory=False)

# Rellenar vacíos con 0
df = df.fillna(0)

# Convertir columnas numéricas
for col in df.columns[7:]:  # todas las de dosis son numéricas
    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

print(f"  {len(df)} filas cargadas")
print("  Columnas relevantes:")
cols = df.columns.tolist()
for i, c in enumerate(cols):
    print(f"    [{i}] {c}")

# ─── IDENTIFICAR COLUMNAS POR GRUPO DE EDAD ────────────────────────────────
# Estructura del CSV (basada en la cabecera):
# SRP PRIMERA:
#   col 7:  SRP 6 A 11 MESES PRIMERA
#   col 8:  SRP 1 ANIO  PRIMERA
#   col 9:  SRP 2 A 5 ANIOS PRIMERA
#   col 10: SRP 6 ANIOS PRIMERA
#   col 11: SRP 7 A 9 ANIOS PRIMERA
#   col 12: SRP 10 A 12 ANIOS PRIMERA
#   col 13: SRP 13 A 19 ANIOS PRIMERA
#   col 14: SRP 10 A 19 ANIOS PRIMERA
#   col 15: SRP 20 A 29 ANIOS PRIMERA
#   col 16: SRP 30 A 39 ANIOS PRIMERA
#   col 17: SRP 40 A 49 ANIOS PRIMERA
#   ...
# SRP SEGUNDA:
#   col 22: SRP 18 MESES SEGUNDA
#   col 23: SRP 2 A 5 ANIOS SEGUNDA
#   col 24: SRP 6 ANIOS SEGUNDA
#   col 25: SRP 7 A 9 ANIOS SEGUNDA
#   col 26: SRP 10 A 12 ANIOS SEGUNDA
#   col 27: SRP 13 A 19 ANIOS SEGUNDA
#   col 28: SRP 10 A 19 ANIOS SEGUNDA
#   col 29: SRP 20 A 29 ANIOS SEGUNDA
#   col 30: SRP 30 A 39 ANIOS SEGUNDA
#   col 31: SRP 40 A 49 ANIOS SEGUNDA
# SR PRIMERA: cols 32-44
# SR SEGUNDA: cols 45-57

# Función para buscar columnas por palabra clave
def find_cols(df, keywords):
    """Devuelve lista de columnas que contienen TODAS las keywords."""
    result = []
    for col in df.columns:
        col_up = col.upper()
        if all(kw.upper() in col_up for kw in keywords):
            result.append(col)
    return result

# ─── MAPEO DE GRUPOS ──────────────────────────────────────────────────────────
# Para cada grupo de edad, sumar SRP+SR, primera+segunda dosis

def sum_group(df, col_lists):
    """Suma múltiples listas de columnas."""
    total = pd.Series(0, index=df.index)
    for cols_list in col_lists:
        for c in cols_list:
            total += pd.to_numeric(df[c], errors='coerce').fillna(0)
    return int(total.sum())

# 6 a 11 meses (solo primera dosis, SRP y SR)
cols_6_11_srp1 = find_cols(df, ['SRP', '6 A 11', 'PRIMERA'])
cols_6_11_sr1  = find_cols(df, ['SR',  '6 A 11', 'PRIMERA'])
# No hay segunda dosis para 6-11 meses
dosis_6_11 = sum_group(df, [cols_6_11_srp1, cols_6_11_sr1])

# 1 año (primera dosis SRP y SR)
cols_1a_srp1 = find_cols(df, ['SRP', '1 ANIO', 'PRIMERA'])
cols_1a_sr1  = find_cols(df, ['SR',  '1 ANIO', 'PRIMERA'])
dosis_1anio  = sum_group(df, [cols_1a_srp1, cols_1a_sr1])

# 18 meses (segunda dosis SRP y SR)
cols_18m_srp2 = find_cols(df, ['SRP', '18 MESES', 'SEGUNDA'])
cols_18m_sr2  = find_cols(df, ['SR',  '18 MESES', 'SEGUNDA'])
dosis_18m     = sum_group(df, [cols_18m_srp2, cols_18m_sr2])

# Rezagados 2 a 12 años:
# SRP: 2A5 1ra+2da, 6ANIOS 1ra+2da, 7A9 1ra+2da, 10A12 1ra+2da
# SR:  2A5 1ra+2da, 6ANIOS 1ra+2da, 7A9 1ra+2da, 10A12 1ra+2da
srp_2a5_1   = find_cols(df, ['SRP', '2 A 5', 'PRIMERA'])
srp_2a5_2   = find_cols(df, ['SRP', '2 A 5', 'SEGUNDA'])
srp_6a_1    = find_cols(df, ['SRP', '6 ANIOS', 'PRIMERA'])
srp_6a_2    = find_cols(df, ['SRP', '6 ANIOS', 'SEGUNDA'])
srp_7a9_1   = find_cols(df, ['SRP', '7 A 9', 'PRIMERA'])
srp_7a9_2   = find_cols(df, ['SRP', '7 A 9', 'SEGUNDA'])
srp_10a12_1 = find_cols(df, ['SRP', '10 A 12', 'PRIMERA'])
srp_10a12_2 = find_cols(df, ['SRP', '10 A 12', 'SEGUNDA'])
sr_2a5_1    = find_cols(df, ['SR',  '2 A 5', 'PRIMERA'])
sr_2a5_2    = find_cols(df, ['SR',  '2 A 5', 'SEGUNDA'])
sr_6a_1     = find_cols(df, ['SR',  '6 ANIOS', 'PRIMERA'])
sr_6a_2     = find_cols(df, ['SR',  '6 ANIOS', 'SEGUNDA'])
sr_7a9_1    = find_cols(df, ['SR',  '7 A 9', 'PRIMERA'])
sr_7a9_2    = find_cols(df, ['SR',  '7 A 9', 'SEGUNDA'])
sr_10a12_1  = find_cols(df, ['SR',  '10 A 12', 'PRIMERA'])
sr_10a12_2  = find_cols(df, ['SR',  '10 A 12', 'SEGUNDA'])
dosis_2_12  = sum_group(df, [
    srp_2a5_1, srp_2a5_2, srp_6a_1, srp_6a_2,
    srp_7a9_1, srp_7a9_2, srp_10a12_1, srp_10a12_2,
    sr_2a5_1, sr_2a5_2, sr_6a_1, sr_6a_2,
    sr_7a9_1, sr_7a9_2, sr_10a12_1, sr_10a12_2
])

# 13 a 19 años (SRP y SR, primera + segunda)
# Nota: el CSV tiene "SRP 13 A 19" y "SRP 10 A 19" (que incluye 10-19)
# Solo usamos la columna "13 A 19" específica para evitar doble conteo con 10-12
srp_13a19_1 = find_cols(df, ['SRP', '13 A 19', 'PRIMERA'])
srp_13a19_2 = find_cols(df, ['SRP', '13 A 19', 'SEGUNDA'])
sr_13a19_1  = find_cols(df, ['SR',  '13 A 19', 'PRIMERA'])
sr_13a19_2  = find_cols(df, ['SR',  '13 A 19', 'SEGUNDA'])
dosis_13_19 = sum_group(df, [srp_13a19_1, srp_13a19_2, sr_13a19_1, sr_13a19_2])

# 20 a 39 años (SRP y SR, primera + segunda)
srp_20a29_1 = find_cols(df, ['SRP', '20 A 29', 'PRIMERA'])
srp_20a29_2 = find_cols(df, ['SRP', '20 A 29', 'SEGUNDA'])
srp_30a39_1 = find_cols(df, ['SRP', '30 A 39', 'PRIMERA'])
srp_30a39_2 = find_cols(df, ['SRP', '30 A 39', 'SEGUNDA'])
sr_20a29_1  = find_cols(df, ['SR',  '20 A 29', 'PRIMERA'])
sr_20a29_2  = find_cols(df, ['SR',  '20 A 29', 'SEGUNDA'])
sr_30a39_1  = find_cols(df, ['SR',  '30 A 39', 'PRIMERA'])
sr_30a39_2  = find_cols(df, ['SR',  '30 A 39', 'SEGUNDA'])
dosis_20_39 = sum_group(df, [
    srp_20a29_1, srp_20a29_2, srp_30a39_1, srp_30a39_2,
    sr_20a29_1, sr_20a29_2, sr_30a39_1, sr_30a39_2
])

# 40 a 49 años (SRP y SR, primera + segunda)
srp_40a49_1 = find_cols(df, ['SRP', '40 A 49', 'PRIMERA'])
srp_40a49_2 = find_cols(df, ['SRP', '40 A 49', 'SEGUNDA'])
sr_40a49_1  = find_cols(df, ['SR',  '40 A 49', 'PRIMERA'])
sr_40a49_2  = find_cols(df, ['SR',  '40 A 49', 'SEGUNDA'])
dosis_40_49 = sum_group(df, [srp_40a49_1, srp_40a49_2, sr_40a49_1, sr_40a49_2])

# ─── MOSTRAR RESULTADOS ───────────────────────────────────────────────────────
print("\n=== DOSIS APLICADAS POR GRUPO DE EDAD (SRP + SR) ===")
grupos = {
    '6 a 11 meses':        dosis_6_11,
    '1 año':               dosis_1anio,
    '18 meses':            dosis_18m,
    'Rezagados 2-12 años': dosis_2_12,
    '13 a 19 años':        dosis_13_19,
    '20 a 39 años':        dosis_20_39,
    '40 a 49 años':        dosis_40_49,
}
for g, d in grupos.items():
    print(f"  {g:<25} {d:>8,}")

total_dosis = sum(grupos.values())
print(f"\n  {'TOTAL':<25} {total_dosis:>8,}")

# ─── ACTUALIZAR EL EXCEL ──────────────────────────────────────────────────────
print("\nActualizando Excel...")
wb = openpyxl.load_workbook(XLSX_SRC)
ws = wb.active

# Las filas de datos comienzan en la fila 4 (índice 1-based)
# Fila 4: 6 a 11 meses
# Fila 5: 1 año
# Fila 6: 18 meses
# Fila 7: Rezagados 2 a 12 años
# Fila 8: 13 a 19 años
# Fila 9: 20 a 39 años
# Fila 10: 40 a 49 años
# Fila 11: TOTAL

# Columnas: A=Grupo, B=Población 2026, C=Población Meta, D=Dosis Aplicadas, E=Pob. Susceptible, F=Cobertura (%)
fila_dosis = {
    4:  dosis_6_11,
    5:  dosis_1anio,
    6:  dosis_18m,
    7:  dosis_2_12,
    8:  dosis_13_19,
    9:  dosis_20_39,
    10: dosis_40_49,
}

for row_num, dosis in fila_dosis.items():
    # D = Dosis Aplicadas (columna 4)
    ws.cell(row=row_num, column=4).value = dosis
    # C = Población Meta (columna 3)
    pob_meta = ws.cell(row=row_num, column=3).value or 0
    # E = Población Susceptible = Meta - Dosis
    susceptible = max(0, pob_meta - dosis)
    ws.cell(row=row_num, column=5).value = susceptible
    # F = Cobertura (%)
    if pob_meta and pob_meta > 0:
        cobertura = round((dosis / pob_meta) * 100, 2)
    else:
        cobertura = 0.0
    ws.cell(row=row_num, column=6).value = cobertura

# Actualizar TOTAL (fila 11)
total_meta = sum(
    (ws.cell(row=r, column=3).value or 0) for r in range(4, 11)
)
total_susceptible = sum(
    (ws.cell(row=r, column=5).value or 0) for r in range(4, 11)
)
total_cobertura = round((total_dosis / total_meta) * 100, 2) if total_meta else 0.0

ws.cell(row=11, column=4).value = total_dosis
ws.cell(row=11, column=5).value = total_susceptible
ws.cell(row=11, column=6).value = total_cobertura

# Actualizar fuente en fila 2
fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M')
nombre_csv = os.path.basename(CSV_PATH)
ws.cell(row=2, column=1).value = (
    f"Fuente dosis aplicadas: {nombre_csv} (SRP-SR SSA)  |  "
    f"Actualizado: {fecha_actual}  |  Metodología: Cobertura (%) = Dosis Aplicadas / Población Meta × 100"
)

# Guardar
wb.save(XLSX_OUT)
print(f"  ✓ Excel guardado en: {XLSX_OUT}")

# ─── RESUMEN FINAL ────────────────────────────────────────────────────────────
print("\n=== RESUMEN DE COBERTURA ===")
print(f"{'Grupo de Edad':<25} {'Dosis':>8} {'Meta':>8} {'Cobertura':>10}")
print("-" * 55)
labels = ['6 a 11 meses', '1 año', '18 meses', 'Rezagados 2-12 años',
          '13 a 19 años', '20 a 39 años', '40 a 49 años']
for i, (label, dosis) in enumerate(zip(labels, fila_dosis.values())):
    row_num = i + 4
    meta = ws.cell(row=row_num, column=3).value or 0
    cob  = ws.cell(row=row_num, column=6).value or 0
    print(f"{label:<25} {dosis:>8,} {meta:>8,} {cob:>9.2f}%")
print("-" * 55)
print(f"{'TOTAL':<25} {total_dosis:>8,} {total_meta:>8,} {total_cobertura:>9.2f}%")
print("\n✅ Proceso completado exitosamente.")

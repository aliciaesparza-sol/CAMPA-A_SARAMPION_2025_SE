
# -*- coding: utf-8 -*-
"""
Agrega columna K: Poblacion Susceptible Real (metodologia CENSIA)
y un parrafo explicativo al pie del Excel.
"""

import openpyxl, glob, shutil
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── RUTAS ────────────────────────────────────────────────────────────────────
files = glob.glob(r'c:\Users\aicil\OneDrive\ANTIGRAVITY\Anexo*.xlsx')
CENSIA_PATH = files[0]
SRC  = r'c:\Users\aicil\OneDrive\ANTIGRAVITY\Cobertura_SRP-SR_SSA_Durango_2026_CENSIA.xlsx'
DEST = r'c:\Users\aicil\OneDrive\ANTIGRAVITY\Cobertura_SRP-SR_SSA_Durango_2026_FINAL.xlsx'
shutil.copy2(SRC, DEST)

# ─── LEER SUSCEPTIBLES DE DURANGO (Anexo, hoja Susceptibles) ─────────────────
wb_c = openpyxl.load_workbook(CENSIA_PATH, data_only=True)
ws_sus = wb_c['Susceptibles']
ws_meta = wb_c['Pob META']

susc = {}
meta_c = {}
for row in ws_sus.iter_rows(min_row=4, max_row=36, values_only=True):
    if row[0] and 'DURANGO' in str(row[0]).upper():
        # cols: 0=GRUPO | 1=6-11m | 4=1año | 7=18m | 10=Rez2-12 | 13=13-19 | 16=20-39 | 19=40-49 | 22=TOTAL
        susc['6_11m']    = int(row[1]  or 0)
        susc['1anio']    = int(row[4]  or 0)
        susc['18m']      = int(row[7]  or 0)
        susc['rez2_12']  = int(row[10] or 0)
        susc['13_19']    = int(row[13] or 0)
        susc['20_39']    = int(row[16] or 0)
        susc['40_49']    = int(row[19] or 0)
        susc['TOTAL']    = int(row[22] or 0) if isinstance(row[22], (int,float)) else sum([
                            susc['6_11m'],susc['1anio'],susc['18m'],
                            susc['rez2_12'],susc['13_19'],susc['20_39'],susc['40_49']])
        break

for row in ws_meta.iter_rows(min_row=3, max_row=35, values_only=True):
    if row[0] and 'DURANGO' in str(row[0]).upper():
        meta_c['6_11m']   = int(row[1] or 0)
        meta_c['1anio']   = int(row[2] or 0)
        meta_c['18m']     = int(row[3] or 0)
        meta_c['rez2_12'] = int(row[4] or 0)
        meta_c['13_19']   = int(row[5] or 0)
        meta_c['20_39']   = int(row[6] or 0)
        meta_c['40_49']   = int(row[7] or 0)
        break

print("Susceptibles CENSIA Durango:", susc)
print("Meta CENSIA Durango:", meta_c)

# ─── ESTILOS ──────────────────────────────────────────────────────────────────
thin = Side(style='thin', color='BFBFBF')
med  = Side(style='medium', color='8B1A4A')
border_norm = Border(top=thin, left=thin, right=thin, bottom=thin)
border_top  = Border(top=med,  left=thin, right=thin, bottom=thin)

HEADER_FILL  = PatternFill(fgColor='4A0D24', fill_type='solid')   # vino muy oscuro
HEADER_FONT  = Font(color='FFFFFF', bold=True, size=10)
SUB_HDR_FILL = PatternFill(fgColor='F3E8F0', fill_type='solid')   # lila muy claro

RED_FILL    = PatternFill(fgColor='FFB3B3', fill_type='solid')
ORANGE_FILL = PatternFill(fgColor='FFD9A0', fill_type='solid')
GREEN_FILL  = PatternFill(fgColor='B7E4C7', fill_type='solid')
TOTAL_FILL  = PatternFill(fgColor='D6C5E0', fill_type='solid')
NOTE_FILL   = PatternFill(fgColor='F5F5F5', fill_type='solid')

# ─── ABRIR EXCEL A MODIFICAR ──────────────────────────────────────────────────
wb = openpyxl.load_workbook(DEST)
ws = wb.active

# Columna K = 11
ws.column_dimensions['K'].width = 22

# ─── ENCABEZADO doble fila ────────────────────────────────────────────────────
# Fila 2 (subtitulo metodologia) - puede que exista, la usamos para el sub-header
c_hdr = ws.cell(row=3, column=11)
c_hdr.value     = 'Susceptibles Reales\n(Metodología CENSIA)'
c_hdr.font      = HEADER_FONT
c_hdr.fill      = HEADER_FILL
c_hdr.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
c_hdr.border    = border_norm
ws.row_dimensions[3].height = 40

# ─── DATOS POR FILA ───────────────────────────────────────────────────────────
fila_clave = {
    4:  susc['6_11m'],
    5:  susc['1anio'],
    6:  susc['18m'],
    7:  susc['rez2_12'],
    8:  susc['13_19'],
    9:  susc['20_39'],
    10: susc['40_49'],
    11: susc['TOTAL'],
}
# Para colorear por susceptibles vs meta (% residual)
meta_por_fila = {
    4:  meta_c['6_11m'],
    5:  meta_c['1anio'],
    6:  meta_c['18m'],
    7:  meta_c['rez2_12'],
    8:  meta_c['13_19'],
    9:  meta_c['20_39'],
    10: meta_c['40_49'],
}

for fila, val in fila_clave.items():
    c = ws.cell(row=fila, column=11)
    c.value       = val
    c.number_format = '#,##0'
    c.alignment   = Alignment(horizontal='center', vertical='center')
    c.border      = border_norm

    if fila == 11:
        c.fill = TOTAL_FILL
        c.font = Font(bold=True)
    else:
        meta = meta_por_fila.get(fila, 0)
        pct_sus = (val / meta * 100) if meta else 0
        # Verde si hay pocos susceptibles (<20%), naranja 20-50%, rojo >50%
        if pct_sus < 20:
            c.fill = GREEN_FILL
        elif pct_sus < 50:
            c.fill = ORANGE_FILL
        else:
            c.fill = RED_FILL

# ─── BLOQUE DE METODOLOGÍA AL PIE ─────────────────────────────────────────────
# Encontrar la última fila con contenido
max_data_row = 11
for r in range(12, 30):
    for col in range(1, 12):
        if ws.cell(row=r, column=col).value is not None:
            max_data_row = r

nota_row = max_data_row + 2

# Título del bloque metodológico
tc = ws.cell(row=nota_row, column=1)
tc.value = 'METODOLOGIA: Calculo de Poblacion Susceptible'
tc.font  = Font(bold=True, size=11, color='FFFFFF')
tc.fill  = HEADER_FILL
tc.alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws.merge_cells(start_row=nota_row, start_column=1, end_row=nota_row, end_column=11)
ws.row_dimensions[nota_row].height = 22

# Párrafo 1: definición
textos = [
    ('¿Que es la Poblacion Susceptible?',
     'Se refiere a las personas que NO han recibido sus dosis de vacuna que les corresponde y que, '
     'por tanto, son susceptibles de contraer sarampion, rubola y parotiditis.'),

    ('Metodologia CENSIA (Anexo 1)',
     'Poblacion Susceptible = Poblacion META 2026  −  Dosis Aplicadas (1 enero 2025 al 6 febrero 2026)\n'
     'Cuando el resultado es negativo (mas dosis que poblacion meta), se registra como 0 (cero).'),

    ('Grupos de edad y sus metas',
     '• 6 a 11 meses: 50% de los menores de 1 anio estimados para 2026 (6,335 personas)\n'
     '• 1 anio: Cohorte de nacidos en 2025 estimados para 2026 (12,255 personas)\n'
     '• 18 meses: Misma cohorte que 1 anio, para la segunda dosis (12,255 personas)\n'
     '• Rezagados 2 a 12 anios: Poblacion pendiente de esquema de 2 a 12 anios (38,054 personas)\n'
     '• 13 a 19 anios: Adolescentes susceptibles (30,841 personas)\n'
     '• 20 a 39 anios: Adultos jovenes (96,450 personas, 50% de la poblacion total del grupo)\n'
     '• 40 a 49 anios: Adultos (40,498 personas, 50% de la poblacion total del grupo)'),

    ('Fuentes',
     '• Dosis SSA (linea verde/azul): SRP-SR-2025_22-02-2026 06-41-08.csv\n'
     '• Dosis CENSIA (linea vino): Anexo 1 - RDA 3er Trimestre 2025 + SIS CENSIA 1 Oct 2025 al 6 Feb 2026\n'
     '  https://siscensia.salud.gob.mx/sarampion_2025/ssa/reporte.php\n'
     '• Poblacion Susceptible Real: Hoja "Susceptibles" del Anexo 1, datos de Durango'),
]

r = nota_row + 1
for titulo, cuerpo in textos:
    # Sub-titulo
    t = ws.cell(row=r, column=1)
    t.value     = titulo
    t.font      = Font(bold=True, size=10, color='4A0D24')
    t.fill      = SUB_HDR_FILL
    t.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
    ws.row_dimensions[r].height = 18
    r += 1

    # Cuerpo de texto
    n_lineas = cuerpo.count('\n') + 1
    cb = ws.cell(row=r, column=1)
    cb.value     = cuerpo
    cb.font      = Font(size=9, color='333333')
    cb.fill      = NOTE_FILL
    cb.alignment = Alignment(horizontal='left', vertical='top',
                             wrap_text=True, indent=1)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
    ws.row_dimensions[r].height = max(15, n_lineas * 14)
    r += 1
    # Espacio entre bloques
    ws.row_dimensions[r].height = 6
    r += 1

wb.save(DEST)
print("Guardado:", DEST)

print("\n=== SUSCEPTIBLES REALES CENSIA DURANGO ===")
labels = ['6-11m','1 anio','18m','Rez 2-12','13-19','20-39','40-49','TOTAL']
vals   = [susc['6_11m'],susc['1anio'],susc['18m'],susc['rez2_12'],
          susc['13_19'],susc['20_39'],susc['40_49'],susc['TOTAL']]
for l,v in zip(labels, vals):
    print(f"  {l:<12} {v:>8,}")

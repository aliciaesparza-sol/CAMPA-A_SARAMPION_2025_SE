
# -*- coding: utf-8 -*-
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font
import os
from datetime import datetime

# ─── RUTAS ────────────────────────────────────────────────────────────────────
CSV_PATH  = r'C:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\CAMPAÑA SARAMPIÓN 10 SEMANAS\SRP-SR-2025_22-02-2026 06-41-08.csv'
XLSX_PATH = r'C:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\CAMPAÑA SARAMPIÓN 10 SEMANAS\Cobertura_SRP-SR_SSA_Durango_2026_PoblaciónSusceptibleFINAL_22_02_2026.xlsx'

# ─── CARGAR DATOS ────────────────────────────────────────────────────────────
print(f"Cargando CSV: {os.path.basename(CSV_PATH)}...")
df = pd.read_csv(CSV_PATH, low_memory=False).fillna(0)

# Columnas numéricas
for col in df.columns[7:]:
    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

def find_cols(keywords):
    return [c for c in df.columns if all(kw.upper() in c.upper() for kw in keywords)]

def sum_group(kw_list):
    total = 0
    for kws in kw_list:
        cols = find_cols(kws)
        total += df[cols].values.sum()
    return int(total)

# ─── CÁLCULOS ────────────────────────────────────────────────────────────────
print("Calculando dosis por grupo de edad...")
sumas = {
    '6 a 11 meses': sum_group([['SRP', '6 A 11', 'PRIMERA'], ['SR', '6 A 11', 'PRIMERA']]),
    '1 año':        sum_group([['SRP', '1 ANIO', 'PRIMERA'], ['SR', '1 ANIO', 'PRIMERA']]),
    '18 meses':     sum_group([['SRP', '18 MESES', 'SEGUNDA'], ['SR', '18 MESES', 'SEGUNDA']]),
    'Rezagados 2 a 12 años': sum_group([
        ['SRP', '2 A 5', '1'], ['SRP', '2 A 5', '2'],
        ['SRP', '6 ANIOS', '1'], ['SRP', '6 ANIOS', '2'],
        ['SRP', '7 A 9', '1'], ['SRP', '7 A 9', '2'],
        ['SRP', '10 A 12', '1'], ['SRP', '10 A 12', '2'],
        ['SR', '2 A 5', '1'], ['SR', '2 A 5', '2'],
        ['SR', '6 ANIOS', '1'], ['SR', '6 ANIOS', '2'],
        ['SR', '7 A 9', '1'], ['SR', '7 A 9', '2'],
        ['SR', '10 A 12', '1'], ['SR', '10 A 12', '2']
    ]),
    '13 a 19 años': sum_group([['SRP', '13 A 19'], ['SR', '13 A 19']]),
    '20 a 39 años': sum_group([
        ['SRP', '20 A 29'], ['SRP', '30 A 39'],
        ['SR', '20 A 29'], ['SR', '30 A 39']
    ]),
    '40 a 49 años': sum_group([['SRP', '40 A 49'], ['SR', '40 A 49']])
}

# ─── ACTUALIZAR EXCEL ─────────────────────────────────────────────────────────
print(f"Abriendo Excel: {os.path.basename(XLSX_PATH)}...")
wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb.active

# Mapeo de filas
filas = {
    '6 a 11 meses': 4,
    '1 año':        5,
    '18 meses':     6,
    'Rezagados 2 a 12 años': 7,
    '13 a 19 años': 8,
    '20 a 39 años': 9,
    '40 a 49 años': 10
}

total_dosis = 0
for grupo, fila in filas.items():
    dosis = sumas[grupo]
    total_dosis += dosis
    
    # C=Meta, D=Dosis, E=Susceptible, F=Cob
    meta = ws.cell(row=fila, column=3).value or 0
    ws.cell(row=fila, column=4).value = dosis
    susc = max(0, int(meta) - dosis)
    ws.cell(row=fila, column=5).value = susc
    cob = (dosis / meta * 100) if meta else 0
    ws.cell(row=fila, column=6).value = round(cob, 2)

# Actualizar TOTAL
ws.cell(row=11, column=4).value = total_dosis
meta_total = sum(ws.cell(row=r, column=3).value or 0 for r in range(4,11))
ws.cell(row=11, column=5).value = max(0, meta_total - total_dosis)
ws.cell(row=11, column=6).value = round((total_dosis / meta_total * 100), 2) if meta_total else 0

# Fuente
ws.cell(row=2, column=1).value = f"Fuente: {os.path.basename(CSV_PATH)} | Actualizado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

wb.save(XLSX_PATH)
print("¡Archivo actualizado y guardado!")
